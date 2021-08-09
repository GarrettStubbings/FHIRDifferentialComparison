# -*- coding: utf-8 -*-
"""
Created on Sun Jun 27 19:49:08 2021

@author: Stubb
"""

import pylab as pl
import pandas as pd
import json
import xmltodict
import os
import pprint
pp = pprint.PrettyPrinter(indent=1)
import string
import xlsxwriter
import openpyxl as pyxl
import re
import copy
import operator

def get_colour_dict(igs):
    """
    Parameters
    ----------
    igs : list of strings
        list of igs to assign colours to.

    Returns
    -------
    dictionary of colours for unique/shared parent/child elements.

    """
    class_colours = {"shared parent": "e0bc8d",
                     "shared child": "ffe9c4",
                     "{} parent".format(igs[0]): "81ad6f",
                     "{} child".format(igs[0]): "a0c991",
                     "{} parent".format(igs[1]): "83aed4",
                     "{} child".format(igs[1]): "b8c7db"
                    }
    return class_colours

def get_folder(base_directory, target_folder):
    """
    Extend a directory to a new folder (if it doesn't exist)
    """
    extension = target_folder.strip('/')
    if extension not in os.listdir(base_directory):
        os.mkdir(base_directory + '/' + extension + '/')
    return base_directory + '/' + extension + '/'


def reformat_xml_element(xml_element, parent_dictionary):
    """
    the xml elements are an ordered dict of ordered dicts (potentially all the
                                                                   way down)
    the to get to the value of a given element is basically []...[]['@value']
    Pull it out into a nice dictionary recursively
    """
    child_dictionary = {}
    for k, v in xml_element.items():
        if type(v) == list:
            continue
        if k == "@url":
            child_dictionary[k] = v
            continue
        if "@value" in v.keys():
            child_dictionary[k] = v["@value"]
        else:
            child_dictionary[k] = reformat_xml_element(v, child_dictionary)
    
    return child_dictionary

def dictionaryify(element_list, file_type = 'json'):
    """
    The goal: Turn the stupid list of elements into a dictionary
    every element in the list has an attribute called ID, transform it into a
    dictionary with the ID as the key.
    """
    if file_type == 'json':
        id_string = "id"
    else:
        id_string = "@id"
    element_dict = {}
    for e in element_list:
        
        name = e[id_string]
        formatted_element = {k:v for k, v in e.items() if id_string not in k}
        
        if file_type == 'json':
            element_dict[name] = formatted_element
        else:
            formatted_element = reformat_xml_element(formatted_element, {})
            element_dict[name] = formatted_element

        
    return element_dict

def get_base_resource(resource_name, base_resource_dir = 
                                                  'Data/HL7Resources/'):
    """
    Most profiles are derivative of a base profile.
    This function grabs the base resource file (the snapshot view)
    
    I've hardcoded some stuff in here - like excluding metadata etc.
    Bad practice I know.
    """
    #if resource_name.lower() not in os.listdir(base_resource_dir):
    #    print(resource_name, "Not in FHIR master?")
    #    return "DNE"
    if "observation" in resource_name.lower():
        resource_name = "observation"
    file_name = base_resource_dir + '{}.profile.json'.format(
                                        resource_name.lower(), resource_name)
    matching_files = [f for f in os.listdir(base_resource_dir) if
                                          resource_name.lower() in f.lower()]
    if len(matching_files) == 0:
        print(resource_name, "Not in FHIR master?")
        return "DNE"
    with open(file_name, encoding="utf8") as f:
        profile_dict = json.load(f)
    element_list = profile_dict['snapshot']['element']
    profile_dict = dictionaryify(element_list, 'json')
    bad_keys = ["id", "implicitRules","contained","modifierExtension","meta"]
    for k in bad_keys:
        bad_key = resource_name + "." + k
        if bad_key in profile_dict.keys():
            del profile_dict[bad_key]

    #pp.pprint(element_list)
    return profile_dict

def split_elements(resource_dicts, base_resource_dict, igs, profile_name,
                   inherit_base):
    """
    This function is now defunct in favour of building resources directly from
    the dictionary as done in higher-level functions.py
    
    Basically it's checking where an element belongs e.g. share/which ig and 
    whether or not it's a child element. But it doesnt do a good job.
    """
    base_resource_element_names = base_resource_dict.keys()
    shared_elements = []
    child_elements = []
    unique_elements = []
    parent_elements = []
    unique_element_igs = []
    child_element_igs = []
    # loop through the elements and check if they're present in both profiles
    element_names = []
    for i, resource_dict in enumerate(resource_dicts):
        element_names.append(list(resource_dict.keys()))

    
    for i in range(len(element_names)):
        parent_element = 'NothingYet'

        for element_name in element_names[i]: 
            if element_name == profile_name:
                continue
            
            # if the name is in the other IG it is shared
            
            if inherit_base[i]:
                shared_condition = (element_name in base_resource_element_names
                                  or element_name in element_names[(i+1)%2])
            else:
                #print("Not Inheriting")
                #print(element_name)
                #pp.pprint(element_names[(i+1)%2])
                shared_condition = element_name in element_names[(i+1)%2]
                
            if shared_condition:
                # have to check whether or not it is a child element
                if parent_element in element_name:
                    child_elements.append(element_name)
                    child_element_igs.append('both')
                    parent_elements.append(parent_element)
                else:
                    shared_elements.append(element_name)
                    parent_element = element_name #.strip(profile_name)
                    
            else:
                # have to check whether or not it is a child element
                if parent_element in element_name:
                    child_elements.append(element_name)
                    if shared_condition:
                        child_element_igs.append('both')
                    else:
                        child_element_igs.append(igs[i])
                    parent_elements.append(parent_element)
                else:
                    unique_elements.append(element_name)
                    unique_element_igs.append(igs[i])
                    parent_element = element_name #.strip(profile_name)
    shared_elements = list(set(shared_elements))
    parent_elements = list(set(parent_elements))
    bad_elements = []
    for i, element in enumerate(shared_elements):
        if element in parent_elements or element in child_elements:
            bad_elements.append(i)
    for i in bad_elements[::-1]:
        shared_elements.pop(i)
        
        
    
    return (shared_elements, child_elements, parent_elements, unique_elements,
                                    unique_element_igs, child_element_igs)

def organize_elements(element_lists, base_resource_dict, IGs, profile_name,
                      inherit_base):
    """ 
    This function is now defunct as well.
    
    Its job was to put all of the elements in the correct order given 
    information like which IG they're from and if they're a parent/child
    """

    (shared_elements, child_elements, parent_elements, unique_elements, 
                unique_element_igs, child_element_igs) = split_elements(
                    element_lists, base_resource_dict, IGs, profile_name,
                    inherit_base)
                    
    ordered_element_list = [profile_name]
    element_classification = ['shared parent']
    # shared elements (shouldn't be duplicates but just in case)
    for i, element in enumerate(shared_elements):
        if element not in ordered_element_list:
            ordered_element_list.append(element)
            element_classification.append('shared child')
    # unique elements (shouldn't be duplicates but just in case)
    for i, element in enumerate(unique_elements):
        if element not in ordered_element_list:
            ordered_element_list.append(element)
            element_classification.append(unique_element_igs[i])
    # parent elements
    for i, parent_element in enumerate(parent_elements):
        ordered_element_list.append(parent_element)
        element_classification.append('shared parent')
        # gotta find ALL of the children
        for j, element in enumerate(child_elements):
            if parent_element in element:
                ordered_element_list.append(element)
                if child_element_igs[j] == 'both':
                    element_classification.append('shared child')
                else:
                    element_classification.append(child_element_igs[j])
    if len(ordered_element_list) > 1:
        if ordered_element_list[0] == ordered_element_list[1].lower():
            ordered_element_list.pop(0)
            element_classification.pop(0)
    return ordered_element_list, element_classification


def add_mappings(element, profile_dicts, base_resource_dict,
                    element_classifications):
    values = []
    for i, profile in enumerate(profile_dicts):
        value = ''
        # if the element is in the profile then we fill the relevant inforation
        if element in profile.keys():
            if 'mapping' in profile[element].keys():
                #pp.pprint(profile[element]['mapping'])
                if type(profile[element]['mapping']) == list:
                    viable_mappings = [d['map'] for d in
                                       profile[element]['mapping'] if
                                       'CDSS5.1' in d['identity']]
                    for i, m in enumerate(viable_mappings):
                        value += " " + m + " "
                        if i < len(viable_mappings) - 1:
                            value += '|'
                else:
                    if 'CDS' in profile[element]['mapping']['identity']:
                        value = profile[element]['mapping']['map']
        
        values.append(value)
    
    diff_value = ''    
    values.append(diff_value)
    return values

def compare_flags(element, profile_dicts, base_resource_dict,
                    element_classification):
    """
    This function is for comparing flags in the output.
    It returns the respective flag and characterizes the difference
    """
    values = []
    for i, profile in enumerate(profile_dicts):
        value = ''
        # if the element is in the profile then we fill the relevant inforation
        if element in profile.keys():
            if 'isModifier' in profile[element].keys():
                value += '!?'
            if 'mustSupport' in profile[element].keys():
                value += 'S'
        values.append(value)
    
    diff_value = ''
    if len(values[1].replace(values[0], '')) > 0:
        diff_value = 'Strengthened'
    elif len(values[0].replace(values[1], '')) > 0:
        diff_value = 'Relaxed'
    
    # Question of whether or not to talk about MS flags on unique elements,
    # right now it seems like they should stay
    # if element not in base_resource_dict.keys():
    #    diff_value = ''
        
    values.append(diff_value)
    return values

def compare_cardinality(element, profile_dicts, base_resource_dict,
                    element_classification):
    """
    This function is for comparing cardinalities in the output.
    It returns the respective cardinalities and characterizes the difference
    
    Note: it brings in the cardinality of the base resource to fill it in
    """
    has_default = 0
    if element in base_resource_dict.keys():
        base_min = base_resource_dict[element]['min']
        base_max = base_resource_dict[element]['max']
        has_default = 1
    else:
        base_min = '0'
        base_max = '1'
    values = []
    for i, profile in enumerate(profile_dicts):
        value = ''
        use_card = 0
        # if the element is in the profile then we fill the relevant inforation
        if element in profile.keys():
            
            card_min = base_min
            card_max = base_max
            if 'min' in profile[element].keys():
                card_min = profile[element]['min']
                use_card = 1
            if 'max' in profile[element].keys():
                card_max = profile[element]['max']
                use_card = 1
            if use_card or has_default:
                value = '{0}..{1}'.format(card_min, card_max)
            

            
        elif element in base_resource_dict.keys() and (
                "shared" in element_classification):
            value = '{0}..{1}'.format(base_min, base_max)

        values.append(value)
    
    
    # Diff Value: strengthen implies min: 0->1, or max * -> 1
    #             relax implies min: 1 -> 0, max: 1->*
    

    diff_value = ''
    if len(values[0]) != 0 and len(values[1]) != 0:
        for i in range(2):
            # IPS min/max      CAB min/max
            if values[0][i] != values[1][i]:
                if values[0][i] == '1':
                    diff_value = 'Relaxed'
                else:
                    diff_value = 'Strengthened'

    values.append(diff_value)
    return values


def compare_types(element, profile_dicts, base_resource_dict,
                    element_classification):
    """
    This function is for comparing types in the output.
    It returns the respective type and characterizes the difference
    
    Note, this was never really meant to deal with using a snapshot view of 
    the resource. It doesn't expect to have a type unless it's in the diff.
    So if using a snapshot view of e.g. the base resource it will spit out
    default values.
    """
    values = []
    num_present = 0
    for i, profile in enumerate(profile_dicts):
        value = ''
        # if the element is in the profile then we fill the relevant inforation
        if element in profile.keys():
            num_present += 1
            # If the type is in the Dif
            if 'type' in profile[element].keys():
                # if the type only has a name it will just be a dict with code in it
                type_dict = profile[element]['type']
                if type(type_dict) == list:
                    type_dict = type_dict[0]

                value += type_dict['code']
                if 'targetProfile' in type_dict.keys():
                    value += ' | '
                    if type(type_dict['targetProfile']) == list:
                        for url in type_dict['targetProfile']:
                            value += url + " | "
                    else:
                        value += type_dict['targetProfile']
                elif 'profile' in type_dict.keys():
                    if type(type_dict['profile']) == list:
                        for x in type_dict['profile']:
                            value += ' | ' + x
                    else:
                        value += ' | ' + type_dict['profile']

        values.append(value)
    
    diff_value = ''
    # in the case that it's a unique element
    if "shared" not in element_classification:    
        diff_value = ''
    else:
        if len(values[0]) > 0 and len(values[1]) == 0:
            diff_value = 'Relaxed'
        elif len(values[0]) == 0 and len(values[1]) > 0:
            diff_value = 'Strengthened'
        elif len(values[0]) > 0:
            if values[0] != values[1]:
                diff_value = "Changed"
    
    values.append(diff_value)
    return values

def compare_binding(element, profile_dicts, base_resource_dict,
                    element_classification):
    """
    This function is for comparing bindings (strengths and valuesets) in the
    output.
    It returns the respective bindings and characterizes the differences based
    only on the strengths, not the valuesets
    """
    values = []
    for i, profile in enumerate(profile_dicts):
        value = ''
        # if the element is in the profile then we fill the relevant info
        if element in profile.keys():
            # If the binding is in the Dif
            if 'binding' in profile[element].keys():
                value += profile[element]['binding']['strength']
                value += ' | ' + profile[element]['binding']['valueSet']

        values.append(value)
    
    # in the case that it's a unique element
    if element not in base_resource_dict.keys():
        diff_value = ''
    else:
        strengths = [v.split()[0] if len(v) > 0 else '' for v in values]
        #print(strengths)
        hierarchy = pl.asarray(['required', 'extensible', 'preferred',
                                                    'example'])[::-1]
        strengths = [pl.where(hierarchy == s)[0] if s != '' else -1 for s
                                                             in strengths]

        diff_value = ''
        if strengths[0] > strengths[1]:
            diff_value = 'Relaxed'
        elif strengths[0] < strengths[1]:
            diff_value = 'Strengthened'
        

        
    values.append(diff_value)
    return values

def compare_slice_descriptions(element, profile_dicts, base_resource_dict,
                    element_classification):
    """
    This function is for comparing slicing in the output.
    It returns the respective slicing and characterizes the difference
    
    I didn't include a value judgement on the differential, so mostly it
    just says changed
    """
    values = []
    for i, profile in enumerate(profile_dicts):
        value = ''
        # if the element is in the profile then we fill the relevant info
        if element in profile.keys():
            # If the slicing is in the Dif
            if 'slicing' in profile[element].keys():
                if 'rules' in profile[element]['slicing'].keys():
                    value += profile[element]['slicing']['rules']
                discriminator_dict = profile[element]['slicing'][
                                                            'discriminator']
                if type(discriminator_dict) == list:
                    discriminator_dict = discriminator_dict[0]
                value += ' by ' + discriminator_dict['type']
                value += ':' + discriminator_dict['path']

        values.append(value)
    
    diff_value = ''
    if values[0] != values[1]:
        diff_value == 'Changed'
    values.append(diff_value)
    return values

def build_dif_table(ordered_element_names, IGs, profile_dicts,
                base_resource_dict, element_classifications,
                include_mappings = False, strip_profile_name = False):
    """
    The whole kit and caboodle
    Compare: element type? Must support, Cardinality, type, binding strength,
    binding valueset? slice description? description?
    get element from name, check if the field exists in the dif (if missing
                                                                 leave nan)
    
    MAPPINGS: Optional attribute to check for
    
    FLAGS: Check element for key "mustSupport" and "isModifier", if L present
                            and R not: relaxed, if R but not L: strengthened
    
    Cardinality: Check element for 'min' and 'max', if either is present,
                    resolve missing with base-resource value
                    option is there to always resolve using base resource
                    if Min in L and not R: relaxed. Ifmax == 1 in L but not R:
                        Relaxed
                    
                    Default to always fill with base, compare strings for Diff
    
    Type: Value should be under "code", could have more data in
                                            "targetProfile" and others?
            Question is whether or not to jam it in (jam it for now)
            
    Binding: Dictionary for Binding with key 'binding', has attributes:
                'strength': binding strength (with attached hierarchy)
                'Valueset': Will just be an URL, jam it in there by default
                
    Slice Description: dictionary under 'slicing' key, human readable form is
                                                    'rules' by 'type' : 'path'
    """
    profile_name = ordered_element_names[0]
    
    formatting_functions = [add_mappings]*include_mappings +  [compare_flags,
                                compare_cardinality, compare_types,
                                compare_binding, compare_slice_descriptions]
    attributes = ['Mapping']*include_mappings + ['Flags', 'Cardinality',
                                    'Type', 'Binding', 'Slice Description']
    
    column_names = ['Element'] + len(attributes)*[IGs[0], IGs[1], 'Diff']
    table = [column_names]
    
    for i, element_name in enumerate(ordered_element_names):
        row_values = []
        element_classification = element_classifications[i]
        # strip the profile name out for the readable thing
        row_name = element_name
        if strip_profile_name:
            if element_name != profile_name:
                row_name = element_name.replace(profile_name + ".", "")
            else:
                row_name = element_name
        
        
        row_values.append(row_name)
        
        # now go through the elements:
        for j, a in enumerate(attributes):
            formatting_function = formatting_functions[j]
            values = formatting_function(element_name, profile_dicts,
                                         base_resource_dict,
                                         element_classification)
            row_values += values
        table.append(row_values)
    if len(table) > 2:
        if table[1][0] == table[2][0].lower():
            table.pop(1)
    return pd.DataFrame(table)

def as_text(value):
    """
    for grabbing values from an excel file without breaking it
    """
    if value is None:
        return ""
    return str(value)

def count_lines(string_input, word = '\n'):
    """
    for finding the height of an excel cell, never ended up being used
    """
    count = sum(1 for _ in re.finditer(r'\b%s\b' % re.escape(word),
                                                       string_input))
    return count

def get_cell_colour(colour):
    """
    for colouring excel cells
    """
    return pyxl.styles.PatternFill(start_color=colour,
                                    end_color=colour, fill_type = "solid")
    

def get_coloured_border(colour):
    """
    for getting borders of excel cells
    """
    border = pyxl.styles.Border(left= pyxl.styles.Side(border_style='thin',
                                                       color=colour),
                    right= pyxl.styles.Side(border_style='thin',
                                            color=colour),
                    top= pyxl.styles.Side(border_style='thin',
                                          color=colour),
                    bottom= pyxl.styles.Side(border_style='thin',
                                             color=colour))
    return border

def get_left_border(colour, other_colour):
    border = pyxl.styles.Border(left= pyxl.styles.Side(border_style='thin',
                                                       color=colour),
                    right= pyxl.styles.Side(border_style='thin',
                                            color=other_colour),
                    top= pyxl.styles.Side(border_style='thin',
                                          color=other_colour),
                    bottom= pyxl.styles.Side(border_style='thin',
                                             color=other_colour))
    return border

def export_excel(table, file_name, attributes):
    """
    Exports the tabular data to an excel spreadsheet
    """
    
    writer = pd.ExcelWriter(file_name, engine='xlsxwriter')
    table.to_excel(writer, index=False, header=False, startrow=1,
                                                       sheet_name = 'Sheet1')
    num_rows, num_cols = table.shape
    num_cols -= 1
    workbook  = writer.book
    worksheet = writer.sheets['Sheet1']
    
    # Add attribute labels: Row 1 merge (b-d), (e-g), etc basiac
    alphabet = string.ascii_uppercase

    offset = 1
    width = 3
    row = 1
    
    merge_format = workbook.add_format({
                                        'bold': 1,
                                        'border': 1,
                                        'align': 'center',
                                        'valign': 'vcenter'})
    
    for i, attribute in enumerate(attributes):
        start_col = offset + i*width
        end_col = start_col + width - 1
        positions = "{0}{2}:{1}{2}".format(alphabet[start_col],
                                                       alphabet[end_col], row)
        

        worksheet.merge_range(positions, attribute, merge_format)

    writer.save()
    
def format_excel(file_name, element_classifications, table, igs, attributes):
    """
    Parameters
    ----------
    file_name : string
        file name of the excel file to EDIT: Not for generating.
    element_classifications : list of strings
        classifies the element as shared/unique parent/child.
    table : Pandas dataframe
        Tabular form of output
    igs : list of strings
        names of igs.
    attributes : List of Strings
        Which attributes of the fhir resources to compare.

    Returns
    -------
    Nothing, just edits an excel file.

    """
    
    # old colours 
    """
    colours = ["a0c991", "b8c7db"]
    class_colours = {"shared parent": "e39922",
                     "shared child": "ffe9c4",
                     igs[0]: colours[0],
                     igs[1]: colours[1]
                        }
    """
    class_colours = get_colour_dict(igs)
    num_rows, num_cols = table.shape
    

    white = "ffffff"
    black = "000000"
    workbook = pyxl.load_workbook(filename = file_name)
    worksheet = workbook.active
    alphabet = string.ascii_uppercase
    
    # add the colours to the IPS and CA-Basline Column Headers
    offset = 1
    width = 3
    row = 2
    header_colours = [class_colours[ig + " child"] for ig in igs]
    for i, attribute in enumerate(attributes):
        start_col = offset + i*width
        position = "{0}{1}".format(alphabet[start_col + width - 1], row)
        worksheet[position].font = pyxl.styles.Font(color=white)
        worksheet[position].fill = pyxl.styles.PatternFill(start_color=black,
                                        end_color=black, fill_type = "solid")
        for n, c in enumerate(header_colours):
            position = "{0}{1}".format(alphabet[start_col + n], row)
            worksheet[position].fill = pyxl.styles.PatternFill(start_color=c,
                                            end_color=c, fill_type = "solid")
    
    # setting the default border (white)
    bottom_right_position = "{0}{1}".format(alphabet[num_cols-1], num_rows + 1)
    default_border = get_coloured_border(white)
    attribute_border = get_left_border(black, white)
    
    for row in worksheet['A2:{}'.format(bottom_right_position)]:
        for cell in row:
            # if it's the first column starting the attribute we set the left
            # border to be black
            if cell.column > 1  and (cell.column + 1) %3 == 0:
                cell.border = attribute_border
            else:
                cell.border = default_border
            
    # setting the colors: Gotta know where each element is from unfortunately
    # adding the hyperlink stuff to this point as well: we just pick the very
    # first link.
    for i, element_class in enumerate(element_classifications):
        color = class_colours[element_class]
        row_number = i + 3
        positions = 'A{0}:{1}{0}'.format(row_number, bottom_right_position[0])
        for row in worksheet[positions]:
            for cell in row:
                cell.fill = pyxl.styles.PatternFill(start_color=color,
                                        end_color=color, fill_type = "solid")
                
                # check for hyperlink 
                if cell.value != None:
                    if 'http' in cell.value:
                        if len(cell.value.split()) > 1:
                            link = cell.value.split()[2]
                            cell.hyperlink = link
    
    # format the width of the columns
    # mostly only care about the first column?
    data = table.values
    headers = table.values[0,:]
    
    # diffs are 1 + 3i
    for i, header in enumerate(headers):
        if i > 0:
            attribute = attributes[int((i-1)/width)]
        if i == 0:
            length = pl.amax([len(e) for e in data[:,0]]) + 1
        elif i%3 == 0:
                length = 13
        elif attribute == 'Flags' or attribute == 'Cardinality':
            length = 6 #len(header) + 2
        elif attribute == 'Mapping':
            length = 10
        else:
            length = 13
        worksheet.column_dimensions[alphabet[i]].width = length
    
    # format the height of the rows: Not sure I want to do this yet,
    # newline formatting doesn't really work
    """
    default_height = 13
    for i in range(pl.shape(data)[1]):
        row_index = i + 2
        values = data[i,:]
        row_height = default_height
        row_height += pl.amax([default_height * count_lines(value) for value
                                                                   in values])
        worksheet.row_dimensions[row_index].height = row_height
    """
    

    
    # add a legend
    black_border = get_coloured_border(black)
    worksheet['A1'] = 'Legend to Right'
    worksheet['A1'].border = black_border
    worksheet['A2'].border = black_border
    column = alphabet[num_cols + 1]
    worksheet['{}1'.format(column)] = 'LEGEND'
    worksheet['{}1'.format(column)].border = black_border
    worksheet.column_dimensions[column].width = 20
    index = 2
    for k, v in class_colours.items():
        #print(k, v)
        cell = column + str(index)
        worksheet[cell] = k
        worksheet[cell].fill = pyxl.styles.PatternFill(start_color=v,
                                    end_color=v, fill_type = "solid")
        worksheet[cell].border = black_border
        index += 1
    
    
    # freeze the element names and headers
    worksheet.freeze_panes = 'B3'
    
    workbook.save(file_name)
    
def get_resource_dictionary(profile_name, dummy_ig_full_names, dummy_igs,
                            data_directories, base_resource_dict, preambles,
                            reduce_cds=True, views=["Diff", "Diff"]):
    """
    Parameters
    ----------
    profile_name : string
        name of the profile in question.
    dummy_ig_full_names : list of strings
        basically file extensions versions of the name
    dummy_igs : list of strings
        readable version of the ig names.
    data_directories : list of strings
        paths to get to the data
    base_resource_dict: dictionary of fhir elementdictionaries
        pulled from fhir base resources (unless otherwise specified)
    reduce cds : Boolean
        Whether or not to reduce the cds-s to only include elements with DE. 
        (This is the right thing to do, the rest does not exist)
    views : list of strings (must be Diff or Snapshot)
        Diff will only pull elements out of the diff to compare
        Snapshot will start the dictionary as the base resource elements, then
            update and add elements from the diff
        
    Returns
    -------
    dictionary versions of the resources
    """
    if type(data_directories) != list:
        #print("IG List is not a list, it is:", dummy_igs)
        data_directories = [data_directories]
        dummy_igs = [dummy_igs]
        dummy_ig_full_names = [dummy_ig_full_names]
        views = [views]
        

    profile_dicts = []
    for i, data_dir in enumerate(data_directories):
        preamble = preambles[i]
        file_list = os.listdir(data_dir)
        profile_files = [f for f in file_list if 'structure' in f.lower() and
                         (profile_name + preamble).lower() in f.lower()]
        if 'omd' in dummy_igs[i].lower():
            profile_files = [f for f in file_list if 'profile' in f.lower() and
                         (profile_name + preamble).lower() in f.lower()]
            
        # New strategy is to allow for unique resources
        # Replace the comparison with the Base Resource
        if len(profile_files) == 0:
            print("No Profile for", profile_name, "in", dummy_igs[i])
            if base_resource_dict == "DNE":
                profile_dicts.append({})
                dummy_igs[i] = 'Nothing'
                dummy_ig_full_names[i] = 'Nothing'
                print("Sending Empty Profile")
                continue
            else:
                profile_dicts.append(base_resource_dict)
                dummy_igs[i] = 'fhir-master'
                dummy_ig_full_names[i] = 'FHIR-Base'
                continue
        else:
            profile_file = profile_files[0]
            
        # Load the profile and get it into nice dictionary formatting
        if "json" in profile_file:
            with open(data_dir + profile_file, encoding="utf8") as f:
                profile_dict = json.load(f)
            element_list = profile_dict['differential']['element']
            profile_dict = dictionaryify(element_list, 'json')
        # The xml one is a little wobbly
        else:
            with open(data_dir + profile_file, encoding="utf8") as f:
                profile_dict = xmltodict.parse(f.read())
            profile_dict = dict(profile_dict['StructureDefinition'])
            element_list = profile_dict['differential']['element']
            profile_dict = dictionaryify(element_list, 'xml')
        
        # Reduce the CDS if it comes down to it (IT SHOULD)
        if 'OMD-CDS-S' in dummy_ig_full_names[i] and reduce_cds:
            bad_keys = []
            for k, v in profile_dict.items():
                if type(v) == dict:
                    if "mapping" not in v.keys():
                        bad_keys.append(k)
                    else:
                        if type(v['mapping']) == list:
                            num_mappings = len([d for d in v['mapping']
                                                if "CDS" in d["identity"]])
                        else:
                            num_mappings = int("map" in
                                               v['mapping'].keys())
                        if num_mappings < 1:
                            bad_keys.append(k)
            for k in bad_keys:
                del profile_dict[k]
        # Append it to the list if it's just the diff view
        if views[i].lower() == 'diff':
            profile_dicts.append(profile_dict)
        # if looking for a snapshot we need to do a bit of work
        elif views[i].lower() == "snapshot":
            # print("Doing the snapshot for", dummy_igs[i])
            if base_resource_dict == "DNE":
                # print("Well I would if the base resource existed")
                combined_dict = profile_dict
            else:
                combined_dict = base_resource_dict.copy()
                for k, v in profile_dict.items():
                    combined_dict[k] = v
            profile_dicts.append(combined_dict)
        else:
            print(views[i], "is not a valid View")
            
            

    """
    print(len(profile_dicts))
    if len(profile_dicts) == 1:
        print(type(profile_dicts))
        return profile_dicts[0]
    else:"""
    return profile_dicts
  
    
def create_2_way_diff(profile_name, ig_full_names, igs, data_directories,
                      output_dir, include_mappings, preambles,
                      inherit_base = [True, True], views = ["Diff", "Diff"],
                      reduce_cds=True):
    """
    This function basically wraps up all the above functions to build the diff
    
    However, it is now defunct because other methods work better.
    """
    dummy_igs = pl.copy(igs)
    dummy_ig_full_names = pl.copy(ig_full_names)
    base_resource_dict = get_base_resource(profile_name)
    profile_dicts = get_resource_dictionary(profile_name, dummy_ig_full_names,
                                            dummy_igs, data_directories,
                                            base_resource_dict, preambles,
                                            reduce_cds, views)
    if profile_dicts == None:
        print("No Dice")
        return None
    #pp.pprint(base_resource_dict)
    ordered_element_names, element_classifications = organize_elements(
        profile_dicts, base_resource_dict, dummy_igs, profile_name,
        inherit_base)
    table = build_dif_table(ordered_element_names, dummy_igs, profile_dicts,
                            base_resource_dict, element_classifications, 
                            include_mappings)
    file_name = "{0}{1}.xlsx".format(output_dir, profile_name)
    attributes = ['Mapping']*include_mappings + ['Flags', 'Cardinality',
                                    'Type', 'Binding', 'Slice Description']
    export_excel(table, file_name, attributes)
    format_excel(file_name, element_classifications, table,
                 dummy_ig_full_names, attributes)
    